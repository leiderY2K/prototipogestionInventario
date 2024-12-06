import openpyxl
from openpyxl import Workbook
from codigo_producto import CodigoProducto

class GuardarExcel():
    def __init__(self):
        self.codigo_producto = CodigoProducto()
        self.nombre_archivo = None
    
    def inicializar_excel(self, nombre_archivo="resultados_investigadores.xlsx"):
        self.nombre_archivo = nombre_archivo
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Resultados"
        sheet.append(["INVESTIGADOR", "TÍTULO DEL PRODUCTO", "TIPO DEL PRODUCTO", "AÑO", "NOMBRE DEL ARCHIVO", "link archivo"])
        workbook.save(self.nombre_archivo)
        print(f"Archivo Excel inicializado: {self.nombre_archivo}")

    def agregar_a_excel(self, resultados):
        if not self.nombre_archivo:
            raise ValueError("El archivo Excel no ha sido inicializado.")

        try:
            # Cargar el archivo Excel existente
            workbook = openpyxl.load_workbook(self.nombre_archivo)
            sheet = workbook.active

            # Obtener los códigos ya creados (verificar en la columna del nombre del archivo)
            nombres_archivos_creados = [
                row[4] for row in sheet.iter_rows(min_row=2, values_only=True)
            ]

            for resultado in resultados:
                # Verificar que el resultado tenga todos los elementos necesarios
                if len(resultado) < 6:
                    print(f"Error: Resultado incompleto: {resultado}")
                    continue

                # Desempaquetar todos los elementos necesarios
                investigador = resultado[0] if resultado[0] else ""  # Manejar posibles vacíos
                titulo_producto = resultado[1] if resultado[1] else ""
                tipo_producto = resultado[2] if resultado[2] else ""
                año = resultado[3] if resultado[3] else ""
                nombre_archivo = resultado[4] if resultado[4] else ""
                link_archivo = resultado[5] if resultado[5] else ""

                tipo_producto_final = tipo_producto.split(", ")[0] if tipo_producto else "XCNE"
                
                # Escoger el primer año o 'XANE' si está vacío
                año_final = año.split(", ")[0] if año else "XANE"

                # Generar código único
                codigo_producto = self.codigo_producto.generar_codigo_producto(
                    investigador, año_final, tipo_producto_final
                )
                codigo_final = codigo_producto
                i = 1
                while codigo_final in nombres_archivos_creados:
                    codigo_final = f"{codigo_producto}_{i}"
                    i += 1
                nombres_archivos_creados.append(codigo_final)

                # Crear la fila completa incluyendo el link
                fila = [
                    investigador,
                    titulo_producto,
                    tipo_producto_final,
                    año_final,
                    codigo_final,
                    link_archivo  # Asegurarnos de incluir el link en la fila
                ]

                # Imprimir para depuración
                print(f"Agregando fila: {fila}")
                
                # Agregar la fila al sheet
                sheet.append(fila)

            # Guardar los cambios
            workbook.save(self.nombre_archivo)
            print(f"Datos agregados y archivo Excel guardado como: {self.nombre_archivo}")

        except Exception as e:
            print(f"Error al guardar en Excel: {str(e)}")
            raise